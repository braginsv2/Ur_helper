import psycopg2
import psycopg2.extras
import json
from datetime import datetime
import os
from contextlib import contextmanager
import time
import random
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class DatabaseManager:
    def __init__(self, connection_params=None):
        """Инициализация базы данных PostgreSQL"""
        if connection_params is None:
            # Параметры подключения по умолчанию
            self.connection_params = {
                'host': os.getenv('DB_HOST', 'localhost'),
                'port': os.getenv('DB_PORT', '5432'),
                'database': os.getenv('DB_NAME', 'clients_db'),
                'user': os.getenv('DB_USER', 'postgres'),
                'password': os.getenv('DB_PASSWORD', '1234'),
                'client_encoding': 'UTF8'  # Добавьте эту строку
            }
        else:
            self.connection_params = connection_params
        # Создаем БД если не существует
        self.create_database_if_not_exists()

        # Создаем таблицы
        self.init_database()
    
    @contextmanager
    def get_connection(self):
        """Контекстный менеджер для подключения к базе данных"""
        conn = None
        try:
            conn = psycopg2.connect(**self.connection_params)
            yield conn
        except Exception as e:
            if conn:
                conn.rollback()
            raise e
        finally:
            if conn:
                conn.close()
    
    def init_database(self):
        """Создание таблицы если она не существует"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute('''
                        CREATE SEQUENCE IF NOT EXISTS client_id_seq 
                        START WITH 70044 
                        INCREMENT BY 1 
                        NO MAXVALUE 
                        NO CYCLE
                        ''')
                    cursor.execute('''
                        CREATE SEQUENCE IF NOT EXISTS client_id_seq_24
                        START WITH 24001 
                        INCREMENT BY 1 
                        NO MAXVALUE 
                        NO CYCLE
                        ''')
                    cursor.execute('''
                        CREATE SEQUENCE IF NOT EXISTS client_id_seq_54
                        START WITH 54001 
                        INCREMENT BY 1 
                        NO MAXVALUE 
                        NO CYCLE
                    ''')
                    cursor.execute('''
                    CREATE TABLE IF NOT EXISTS clients (
                        id SERIAL PRIMARY KEY,
                        accident TEXT,           
                        client_id VARCHAR(50) UNIQUE NOT NULL,
                        fio TEXT NOT NULL,
                        seria_pasport TEXT,
                        number_pasport TEXT,
                        where_pasport TEXT,
                        when_pasport TEXT,
                        address TEXT,
                        index_postal TEXT,
                        number TEXT,
                        date_of_birth TEXT,
                        city_birth TEXT,
                        date_dtp TEXT,
                        time_dtp TEXT,
                        address_dtp TEXT,
                        who_dtp TEXT,
                        marks TEXT,
                        car_number TEXT,
                        year_auto TEXT,
                        docs TEXT,
                        dkp TEXT,
                        seria_docs TEXT,
                        number_docs TEXT,
                        data_docs TEXT,
                        insurance TEXT,
                        seria_insurance TEXT,
                        number_insurance TEXT,
                        date_insurance TEXT,
                        fio_culp TEXT,
                        marks_culp TEXT,
                        number_auto_culp TEXT,
                        number_photo TEXT,
                        place TEXT,
                        bank TEXT,
                        bank_account TEXT,
                        bank_account_corr TEXT,
                        "BIK" TEXT,
                        "INN" TEXT,
                        created_at TEXT,
                        data_json TEXT,
                        sobstvenik TEXT,
                        fio_sobs TEXT,
                        date_of_birth_sobs TEXT,
                        answer_ins TEXT,
                        analis_ins TEXT,
                        vibor TEXT,
                        vibor1 TEXT,
                        "Nv_ins" TEXT,
                        date_coin_ins TEXT,
                        "Na_ins" TEXT,
                        "date_Na_ins" TEXT,
                        date_exp TEXT,
                        org_exp TEXT,
                        coin_exp TEXT,
                        date_sto TEXT,
                        time_sto TEXT,
                        address_sto TEXT,
                        coin_exp_izn TEXT,
                        coin_osago TEXT,
                        coin_not TEXT,
                        "N_dov_not" TEXT,
                        data_dov_not TEXT,
                        fio_not TEXT,
                        number_not TEXT,
                        date_ins TEXT,
                        date_pret TEXT,
                        pret TEXT,
                        ombuc Text,
                        data_pret_prin TEXT,
                        data_pret_otv TEXT,
                        "N_pret_prin" TEXT,
                        date_ombuc TEXT,
                        date_ins_pod TEXT,
                        seria_vu_culp TEXT,
                        number_vu_culp TEXT,
                        data_vu_culp TEXT,
                        date_of_birth_culp TEXT,
                        index_culp TEXT,
                        address_culp TEXT,
                        number_culp TEXT,
                        "N_viplat_work" TEXT,
                        date_viplat_work TEXT,
                        "N_plat_por" TEXT,
                        date_plat_por TEXT,
                        sud TEXT,
                        gos_money TEXT,
                        date_izvesh_dtp TEXT,
                        date_isk TEXT,
                        dop_osm TEXT,
                        ev TEXT,
                        address_park TEXT,
                        fio_k TEXT,
                        data_dop_osm TEXT,
                        "viborRem" TEXT,
                        date_zayav_sto TEXT,
                        pret_sto TEXT,
                        data_otkaz_sto TEXT,
                        date_napr_sto TEXT,
                        address_sto_main TEXT,
                        data_sto_main TEXT,
                        time_sto_main TEXT,
                        city_sto TEXT,
                        "Done" TEXT,
                        city TEXT,
                        year TEXT,
                        street TEXT,
                        "N_gui" TEXT,
                        date_gui TEXT,
                        "N_prot" TEXT,
                        date_prot TEXT,
                        date_road TEXT,
                        "N_kv_not" TEXT,
                        date_kv_not TEXT,
                        "N_kv_ur" TEXT,
                        date_kv_ur TEXT,
                        "N_kv_exp" TEXT,
                        status TEXT,
                        fio_c TEXT,
                        fio_c_k TEXT,
                        seria_pasport_c TEXT,
                        number_pasport_c TEXT,
                        where_pasport_c TEXT,
                        when_pasport_c TEXT,
                        address_c TEXT,
                        date_of_birth_c TEXT,
                        coin_c TEXT,
                        city_birth_c TEXT,
                        index_postal_c TEXT,
                        number_c TEXT,
                        money_exp TEXT,
                        user_id TEXT,
                        ur_money TEXT,
                        calculation TEXT
                    )
                    ''')
                    cursor.execute('ALTER TABLE clients ADD COLUMN IF NOT EXISTS agent_id TEXT')
                    cursor.execute('CREATE INDEX IF NOT EXISTS idx_agent_id ON clients(agent_id)')
                    cursor.execute('''
                        CREATE TABLE IF NOT EXISTS admins (
                            id SERIAL PRIMARY KEY,
                            user_id TEXT NOT NULL UNIQUE,
                            fio TEXT NOT NULL,
                            fio_k TEXT,
                            seria_pasport TEXT NOT NULL,
                            number_pasport TEXT NOT NULL,
                            where_pasport TEXT NOT NULL,
                            when_pasport TEXT NOT NULL,
                            date_of_birth TEXT,
                            city_birth TEXT,
                            address TEXT,
                            index_postal TEXT,
                            admin_value TEXT NOT NULL,
                            city_admin TEXT NOT NULL,
                            number TEXT,
                            created_at TEXT,
                            is_active BOOLEAN DEFAULT true,
                            invited_by_user_id TEXT,
                            invited_by_type TEXT
                        )
                    ''')
                    cursor.execute('''
                        ALTER TABLE admins 
                        ADD COLUMN IF NOT EXISTS invited_by_user_id TEXT
                    ''')
                    cursor.execute('ALTER TABLE admins ADD COLUMN IF NOT EXISTS fio_k TEXT')
                    cursor.execute('ALTER TABLE admins ADD COLUMN IF NOT EXISTS date_of_birth TEXT')
                    cursor.execute('ALTER TABLE admins ADD COLUMN IF NOT EXISTS city_birth TEXT')
                    cursor.execute('ALTER TABLE admins ADD COLUMN IF NOT EXISTS address TEXT')
                    cursor.execute('ALTER TABLE admins ADD COLUMN IF NOT EXISTS index_postal TEXT')
                    cursor.execute('''
                        ALTER TABLE admins 
                        ADD COLUMN IF NOT EXISTS invited_by_type TEXT
                    ''')
                    cursor.execute('''
                        ALTER TABLE admins 
                        ADD COLUMN IF NOT EXISTS number TEXT
                    ''')
                    cursor.execute('''
                        CREATE TABLE IF NOT EXISTS agent_contract_stats (
                            id SERIAL PRIMARY KEY,
                            agent_id TEXT NOT NULL,
                            client_user_id TEXT,
                            contract_client_id TEXT NOT NULL,
                            contract_type TEXT NOT NULL,
                            is_first_contract BOOLEAN DEFAULT false,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                        )
                    ''')
                    cursor.execute('CREATE INDEX IF NOT EXISTS idx_agent_contract_stats_agent_id ON agent_contract_stats(agent_id)')
                    cursor.execute('''
                        CREATE TABLE IF NOT EXISTS client_agent_relationships (
                            id SERIAL PRIMARY KEY,
                            client_user_id BIGINT NOT NULL,
                            agent_id BIGINT NOT NULL,
                            client_contract_id VARCHAR(50),
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            UNIQUE(client_user_id)
                        )
                        ''')
                    cursor.execute('''
                        ALTER TABLE client_agent_relationships 
                        ADD COLUMN IF NOT EXISTS client_contract_id VARCHAR(50)
                    ''')
                    cursor.execute('CREATE INDEX IF NOT EXISTS idx_client_agent_relationships_client_id ON client_agent_relationships(client_user_id)')
                    cursor.execute('CREATE INDEX IF NOT EXISTS idx_client_agent_relationships_agent_id ON client_agent_relationships(agent_id)')
                    # Создаем индексы для быстрого поиска
                    cursor.execute('CREATE INDEX IF NOT EXISTS idx_client_id ON clients(client_id)')
                    cursor.execute('CREATE INDEX IF NOT EXISTS idx_fio ON clients(fio)')
                    cursor.execute('CREATE INDEX IF NOT EXISTS idx_fio_gin ON clients USING GIN (to_tsvector(\'russian\', fio))')
                    cursor.execute('CREATE INDEX IF NOT EXISTS idx_admin_user_id ON admins(user_id)')
                    
                    cursor.execute('''
                        CREATE TABLE IF NOT EXISTS pending_approvals (
                            id SERIAL PRIMARY KEY,
                            client_id VARCHAR(50) NOT NULL,
                            user_id TEXT NOT NULL,
                            document_type TEXT NOT NULL, -- 'doverennost' или 'payment'
                            document_url TEXT,
                            fio TEXT NOT NULL,
                            amount DECIMAL(10, 2), -- для payment
                            status TEXT DEFAULT 'pending', -- pending, approved, rejected
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            reviewed_by TEXT,
                            reviewed_at TIMESTAMP
                        )
                    ''')
                    cursor.execute('CREATE INDEX IF NOT EXISTS idx_pending_approvals_status ON pending_approvals(status)')
                    cursor.execute('''
                        ALTER TABLE pending_approvals 
                        ADD COLUMN IF NOT EXISTS rejection_reason TEXT
                    ''')

                    cursor.execute('''
                        ALTER TABLE pending_approvals 
                        ADD COLUMN IF NOT EXISTS receipt_number TEXT
                    ''')

                    cursor.execute('''
                        ALTER TABLE pending_approvals 
                        ADD COLUMN IF NOT EXISTS receipt_uploaded_at TIMESTAMP
                    ''')
                    cursor.execute('''
                        CREATE TABLE IF NOT EXISTS agent_finances (
                            id SERIAL PRIMARY KEY,
                            agent_id TEXT NOT NULL UNIQUE,
                            balance DECIMAL(10, 2) DEFAULT 0,
                            total_earned DECIMAL(10, 2) DEFAULT 0,
                            last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                        )
                    ''')

                    # Таблица для заказов на вывод средств
                    cursor.execute('''
                        CREATE TABLE IF NOT EXISTS withdrawal_requests (
                            id SERIAL PRIMARY KEY,
                            agent_id TEXT NOT NULL,
                            agent_fio TEXT NOT NULL,
                            amount DECIMAL(10, 2) NOT NULL,
                            status TEXT DEFAULT 'pending', -- pending, approved, rejected
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            reviewed_by TEXT,
                            reviewed_at TIMESTAMP
                        )
                    ''')
                    cursor.execute('CREATE INDEX IF NOT EXISTS idx_withdrawal_requests_status ON withdrawal_requests(status)')

                    cursor.execute('''
                        CREATE TABLE IF NOT EXISTS agent_earnings_history (
                            id SERIAL PRIMARY KEY,
                            agent_id TEXT NOT NULL,
                            client_id TEXT NOT NULL,
                            amount DECIMAL(10, 2) NOT NULL,
                            payment_confirmed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                        )
                    ''')
                    cursor.execute('CREATE INDEX IF NOT EXISTS idx_agent_earnings_agent_id ON agent_earnings_history(agent_id)')
                    cursor.execute('CREATE INDEX IF NOT EXISTS idx_agent_earnings_date ON agent_earnings_history(payment_confirmed_at)')


                    conn.commit()
                    
                    print("База данных PostgreSQL инициализирована")
        except Exception as e:
            print(f"Ошибка инициализации базы данных: {e}")
            raise e
    def get_client_contracts(self, user_id):
        """Получение всех договоров клиента по user_id"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    cursor.execute("""
                        SELECT client_id, fio, accident, created_at, status, 
                            COALESCE(data_json, '{}') as data_json
                        FROM clients 
                        WHERE user_id = %s::text 
                        ORDER BY created_at DESC
                    """, (user_id,))
                    
                    results = cursor.fetchall()
                    return [dict(row) for row in results]
        except Exception as e:
            print(f"Ошибка получения договоров клиента: {e}")
            return []
    def create_database_if_not_exists(self):
        """Создание базы данных если она не существует"""
        try:
            # Подключаемся к postgres для создания БД
            temp_params = self.connection_params.copy()
            temp_params['database'] = 'postgres'
            
            conn = psycopg2.connect(**temp_params)
            conn.autocommit = True
            
            with conn.cursor() as cursor:
                # Проверяем существование БД
                cursor.execute("""
                SELECT 1 FROM pg_database WHERE datname = %s
                """, (self.connection_params['database'],))
                
                if not cursor.fetchone():
                    # Создаем БД если не существует
                    cursor.execute(f"CREATE DATABASE {self.connection_params['database']}")
                    print(f"База данных {self.connection_params['database']} создана")
                else:
                    print(f"База данных {self.connection_params['database']} уже существует")
                    
        except Exception as e:
            print(f"Ошибка при создании БД: {e}")
        finally:
            if conn:
                conn.close()

    def save_agent_contract_stat(self, agent_id, client_user_id, contract_client_id, contract_type, is_first_contract=False):
        """Сохранение статистики договора агента"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute('''
                        INSERT INTO agent_contract_stats 
                        (agent_id, client_user_id, contract_client_id, contract_type, is_first_contract)
                        VALUES (%s, %s, %s, %s, %s)
                    ''', (agent_id, client_user_id, contract_client_id, contract_type, is_first_contract))
                    conn.commit()
                    print(f"Сохранена статистика: агент {agent_id}, тип {contract_type}, первый договор: {is_first_contract}")
        except Exception as e:
            print(f"Ошибка сохранения статистики агента: {e}")

    def check_if_first_contract_for_invited_client(self, client_user_id):
        """Проверка, первый ли это договор для приглашенного клиента"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute('''
                        SELECT COUNT(*) FROM agent_contract_stats 
                        WHERE client_user_id = %s AND contract_type = 'invited'
                    ''', (client_user_id,))
                    count = cursor.fetchone()[0]
                    return count == 0
        except Exception as e:
            print(f"Ошибка проверки первого договора: {e}")
            return False

    def get_agent_contract_statistics_detailed(self, agent_id):
        """Получение детальной статистики договоров агента"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    # Прямые договоры (лично заполненные)
                    cursor.execute('''
                        SELECT COUNT(*) FROM agent_contract_stats 
                        WHERE agent_id = %s AND contract_type = 'direct'
                    ''', (agent_id,))
                    direct_contracts = cursor.fetchone()[0]
                    
                    # Первые договоры приглашенных
                    cursor.execute('''
                        SELECT COUNT(*) FROM agent_contract_stats 
                        WHERE agent_id = %s AND contract_type = 'invited' AND is_first_contract = true
                    ''', (agent_id,))
                    first_invited_contracts = cursor.fetchone()[0]
                    
                    # Все договоры приглашенных
                    cursor.execute('''
                        SELECT COUNT(*) FROM agent_contract_stats 
                        WHERE agent_id = %s AND contract_type = 'invited'
                    ''', (agent_id,))
                    all_invited_contracts = cursor.fetchone()[0]
                    
                    return {
                        'direct_contracts': direct_contracts,
                        'first_invited_contracts': first_invited_contracts, 
                        'all_invited_contracts': all_invited_contracts,
                        'total_credited': direct_contracts + first_invited_contracts
                    }
        except Exception as e:
            print(f"Ошибка получения статистики агента: {e}")
            return {'direct_contracts': 0, 'first_invited_contracts': 0, 'all_invited_contracts': 0, 'total_credited': 0}
    def generate_next_client_id(self, city_prefix="70"):
        """Генерация следующего client_id с префиксом города"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    # Определяем имя последовательности по префиксу
                    if city_prefix == "70":
                        seq_name = "client_id_seq"
                    elif city_prefix == "24":
                        seq_name = "client_id_seq_24"
                    elif city_prefix == "54":
                        seq_name = "client_id_seq_54"
                    else:
                        seq_name = "client_id_seq"  # По умолчанию Томск
                    
                    # Проверяем текущее значение последовательности
                    cursor.execute(f"SELECT last_value FROM {seq_name}")
                    seq_value = cursor.fetchone()[0]
                    
                    # Если это первый запуск, синхронизируем с существующими данными
                    min_value = int(city_prefix) * 1000 + 1
                    if seq_value == 1:
                        cursor.execute(f'''
                        SELECT COALESCE(MAX(CAST(client_id AS INTEGER)), {min_value - 1}) 
                        FROM clients 
                        WHERE client_id ~ '^[0-9]+$' 
                        AND client_id LIKE '{city_prefix}%'
                        ''')
                        max_existing = cursor.fetchone()[0]
                        if max_existing >= min_value:
                            cursor.execute(f"SELECT setval('{seq_name}', {max_existing + 1})")
                    
                    # Получаем следующий ID атомарно
                    cursor.execute(f"SELECT nextval('{seq_name}')")
                    next_id = cursor.fetchone()[0]
                    
                    return str(next_id)
        except Exception as e:
            print(f"Ошибка генерации client_id: {e}")
            raise e
    def get_invited_clients_count(self, user_id):
        """Получение количества договоров приглашенных клиентов"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute("""
                        SELECT COUNT(*) 
                        FROM clients c
                        JOIN client_agent_relationships car ON c.user_id::text = car.client_user_id::text
                        WHERE car.agent_id = %s
                    """, (user_id,))
                    
                    result = cursor.fetchone()
                    return result[0] if result else 0
        except Exception as e:
            print(f"Ошибка получения количества приглашенных клиентов: {e}")
            return 0
    def save_client_data_with_generated_id(self, data):
        """Генерирует client_id, добавляет в data и сохраняет в базу ИЛИ обновляет существующего клиента"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    # Текущая дата и время
                    created_at = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
                    
                    # Проверяем, существует ли клиент с таким ФИО
                    client_id_to_check = data.get('client_id', '').strip()
                    cursor.execute("SELECT client_id, fio, data_json FROM clients WHERE client_id = %s", (client_id_to_check,))
                    existing_client = cursor.fetchone()
                    
                    if existing_client:
                        # Клиент существует - обновляем его данные
                        existing_fio = existing_client['fio']
                        existing_client_id = existing_client['client_id']
                        existing_data_json = existing_client['data_json']
                        
                        # Парсим существующие данные из JSON
                        try:
                            existing_data = json.loads(existing_data_json) if existing_data_json else {}
                        except (TypeError, ValueError, json.JSONDecodeError):
                            existing_data = {}
                        
                        # Объединяем существующие данные с новыми (новые имеют приоритет)
                        existing_data_clean = {k: v for k, v in existing_data.items() if k != 'data_json'}
                        merged_data = {**existing_data_clean, **data}
                        merged_data['client_id'] = existing_client['client_id'] # Сохраняем существующий client_id
                        
                        print(f"Обновляем существующего клиента с fio: {existing_fio}")
                        client_user_id = data.get('user_id')
                        if client_user_id and not data.get('agent_id'):
                            cursor.execute(
                                "SELECT agent_id FROM client_agent_relationships WHERE client_user_id = %s",
                                (client_user_id,)
                            )
                            agent_relationship = cursor.fetchone()
                            if agent_relationship:
                                data['agent_id'] = str(agent_relationship['agent_id'])
                        # Подготовка данных для обновления
                        update_data = self._prepare_client_data(merged_data, created_at)
                        
                        # SQL запрос для обновления
                        update_query = '''
                        UPDATE clients SET 
                            accident=%s, client_id=%s, fio=%s, seria_pasport=%s, number_pasport=%s, where_pasport=%s, when_pasport=%s,
                            address=%s, index_postal=%s, number=%s, date_of_birth=%s, city_birth=%s,
                            date_dtp=%s, time_dtp=%s, address_dtp=%s, who_dtp=%s, marks=%s, car_number=%s,
                            year_auto=%s, docs=%s, dkp=%s, seria_docs=%s, number_docs=%s, data_docs=%s,
                            insurance=%s, seria_insurance=%s, number_insurance=%s, date_insurance=%s,
                            fio_culp=%s, marks_culp=%s, number_auto_culp=%s, number_photo=%s, place=%s,
                            bank=%s, bank_account=%s, bank_account_corr=%s, "BIK"=%s, "INN"=%s,
                            created_at=%s, data_json=%s, sobstvenik=%s, fio_sobs=%s, date_of_birth_sobs=%s, answer_ins=%s, analis_ins=%s,
                            vibor=%s, vibor1=%s, "Nv_ins"=%s, date_coin_ins=%s, "Na_ins"=%s, "date_Na_ins"=%s, date_exp=%s, org_exp=%s, coin_exp=%s, 
                            date_sto=%s, time_sto=%s, address_sto=%s, coin_exp_izn=%s, coin_osago=%s, coin_not=%s, 
                            "N_dov_not"=%s, data_dov_not=%s, fio_not=%s, number_not=%s, date_ins=%s, date_pret=%s, pret=%s, ombuc=%s,data_pret_prin=%s,
                            data_pret_otv=%s,"N_pret_prin"=%s, date_ombuc=%s,date_ins_pod=%s, seria_vu_culp=%s, number_vu_culp=%s,data_vu_culp=%s, date_of_birth_culp=%s,
                            index_culp=%s,address_culp=%s,number_culp=%s, "N_viplat_work"=%s, date_viplat_work=%s, "N_plat_por"=%s, date_plat_por=%s, sud=%s, gos_money=%s,
                            date_izvesh_dtp=%s, date_isk=%s, dop_osm=%s, ev=%s, address_park=%s, fio_k=%s, data_dop_osm=%s, "viborRem"=%s, date_zayav_sto=%s, pret_sto=%s, data_otkaz_sto=%s,date_napr_sto=%s,
                            address_sto_main=%s, data_sto_main=%s, time_sto_main=%s, city_sto=%s, "Done"=%s, city=%s, year=%s, street=%s, "N_gui"=%s, date_gui=%s, "N_prot"=%s, date_prot=%s,
                            date_road=%s, "N_kv_not"=%s, date_kv_not=%s, "N_kv_ur"=%s, date_kv_ur=%s, "N_kv_exp"=%s, status=%s, fio_c=%s, fio_c_k=%s, seria_pasport_c=%s,number_pasport_c=%s,
                            where_pasport_c=%s, when_pasport_c=%s, address_c=%s, date_of_birth_c=%s, coin_c=%s, city_birth_c=%s, index_postal_c=%s, number_c=%s, money_exp=%s, user_id = %s,
                            agent_id=%s, ur_money=%s, calculation=%s
                        WHERE client_id=%s
                        '''
                        
                        # Добавляем client_id в конец для WHERE условия
                        update_values = list(update_data.values()) + [existing_client['client_id']]
                        cursor.execute(update_query, update_values)
                        
                        conn.commit()
                        
                        print(f"Клиент обновлен в базе с client_id: {existing_client['client_id']}")
                        
                        # Возвращаем существующий client_id и обновленные данные
                        return existing_client['client_id'], merged_data
                    else:
                        # Клиент с таким client_id не найден
                        print(f"Клиент с client_id '{client_id_to_check}' не найден для обновления")
                        return None, None
        except psycopg2.IntegrityError as e:
            print(f"Конфликт client_id, генерируем новый: {e}")
            return self.save_client_data_with_generated_id(data)
        except Exception as e:
            print(f"Ошибка сохранения клиента: {e}")
            raise e
    def get_city_prefix(self, user_id):
        """Получить префикс client_id на основе города клиента"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(
                        "SELECT city_admin FROM admins WHERE user_id = %s::text",
                        (user_id,)
                    )
                    result = cursor.fetchone()
                    
                    if result and result[0]:
                        city = result[0].strip()
                        
                        # Определяем префикс по городу
                        if city == "Томск":
                            return "70"
                        elif city == "Красноярск":
                            return "24"
                        elif city == "Новосибирск":
                            return "54"
                        else:
                            return "70"  # По умолчанию Томск
                    else:
                        return "70"  # По умолчанию если город не найден
        except Exception as e:
            print(f"Ошибка получения города: {e}")
            return "70"  # По умолчанию при ошибке
    def save_client_data_with_generated_id_new(self, data, max_retries=3):
        """Генерирует client_id, добавляет в data и сохраняет в базу ИЛИ обновляет существующего клиента"""
        for attempt in range(max_retries):
            try:
                with self.get_connection() as conn:
                    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                        # Текущая дата и время
                        created_at = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
                        
                        # Проверяем, существует ли клиент с таким ФИО
                        fio_to_check = data.get('fio', '').strip()
                        cursor.execute("SELECT client_id, data_json FROM clients WHERE fio = %s AND user_id = %s::text", 
                                        (fio_to_check, data.get('user_id')))
                        existing_client = cursor.fetchone()

                        # Клиент не существует - создаем нового
                        print("Создаем нового клиента")
                        
                        # 1. Получаем город клиента для определения префикса
                        client_user_id = data.get('user_id')
                        city_prefix = self.get_city_prefix(client_user_id)

                        # 2. Генерируем client_id с префиксом города
                        client_id = self.generate_next_client_id(city_prefix)

                        # 3. Добавляем client_id в данные
                        data['client_id'] = client_id
                        client_user_id = data.get('user_id')
                        client_user_id = data.get('user_id')
                        creator_user_id = data.get('creator_user_id', client_user_id)  # Кто создает договор

                        if client_user_id and not data.get('agent_id'):
                            cursor.execute(
                                "SELECT agent_id FROM client_agent_relationships WHERE client_user_id = %s",
                                (client_user_id,)
                            )
                            agent_relationship = cursor.fetchone()
                            if agent_relationship:
                                # НЕ записываем agent_id если клиент сам создает
                                if creator_user_id != client_user_id:
                                    data['agent_id'] = str(agent_relationship['agent_id'])
                                data['creator_user_id'] = creator_user_id  # Сохраняем кто создает
                        # Подготовка данных для базы
                        client_data = self._prepare_client_data(data, created_at)
                        
                        # SQL запрос для вставки
                        insert_query = '''
                        INSERT INTO clients (
                            accident, client_id, fio, seria_pasport, number_pasport, where_pasport, 
                            when_pasport, address , index_postal, number , date_of_birth , city_birth ,
                            date_dtp , time_dtp , address_dtp , who_dtp , marks , car_number ,
                            year_auto , docs , dkp , seria_docs , number_docs , data_docs ,insurance , 
                            seria_insurance , number_insurance , date_insurance ,fio_culp , marks_culp , number_auto_culp ,
                            bank , bank_account , bank_account_corr , number_photo , place , "BIK" , "INN" , created_at,
                            data_json , sobstvenik , fio_sobs , date_of_birth_sobs , answer_ins , analis_ins ,
                            vibor , vibor1 , "Nv_ins" , date_coin_ins , "Na_ins" , "date_Na_ins" , 
                            date_exp , org_exp , coin_exp , date_sto , time_sto , address_sto , 
                            coin_exp_izn , coin_osago , coin_not , "N_dov_not" , data_dov_not , fio_not ,
                            number_not , date_ins , date_pret, pret, ombuc, data_pret_prin,
                            data_pret_otv, "N_pret_prin", date_ombuc, date_ins_pod, seria_vu_culp ,number_vu_culp ,
                            data_vu_culp , date_of_birth_culp , index_culp ,address_culp ,number_culp , "N_viplat_work" ,
                            date_viplat_work , "N_plat_por" ,date_plat_por ,sud ,gos_money ,date_izvesh_dtp ,
                            date_isk, dop_osm, ev, address_park , fio_k, data_dop_osm, "viborRem", 
                            date_zayav_sto, pret_sto, data_otkaz_sto, date_napr_sto, address_sto_main, data_sto_main,
                            time_sto_main, city_sto, "Done", city, year, street, 
                            "N_gui", date_gui, "N_prot", date_prot, date_road, "N_kv_not", 
                            date_kv_not, "N_kv_ur", date_kv_ur, "N_kv_exp", status, fio_c,
                            fio_c_k, seria_pasport_c,number_pasport_c, where_pasport_c, when_pasport_c, address_c, 
                            date_of_birth_c, coin_c, city_birth_c, index_postal_c, number_c, money_exp, user_id, agent_id, ur_money, calculation 
                        ) VALUES (  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                    %s, %s
                                    )
                        ON CONFLICT (client_id) DO NOTHING
                        RETURNING id
                        '''
                        print(f"DEBUG: Количество полей в client_data: {len(client_data)}")
                        print(f"DEBUG: Количество плейсхолдеров %s в запросе: {insert_query.count('%s')}")
                        print(f"DEBUG: client_id = {client_id}")
                        print(f"DEBUG: user_id = {data.get('user_id')}")
                        print(f"DEBUG: agent_id = {data.get('agent_id')}")
                        cursor.execute(insert_query, tuple(client_data.values()))
                        result = cursor.fetchone()
                        
                        if result is None:
                            # Конфликт client_id, повторяем попытку
                            conn.rollback()
                            print(f"Конфликт client_id {client_id}, попытка {attempt + 1}")
                            if attempt < max_retries - 1:
                                time.sleep(0.1 + random.uniform(0, 0.1))  # Небольшая задержка
                                continue
                            else:
                                raise Exception("Не удалось создать уникальный client_id после нескольких попыток")
                        
                        db_id = result['id']
                        conn.commit()
                        try:
                            client_user_id = data.get('user_id')
                            creator_user_id = data.get('creator_user_id', client_user_id)  # кто создает договор
                            
                            # Определяем агента и тип договора
                            agent_id, contract_type, is_first = self._determine_contract_stats(client_user_id, creator_user_id)
                            
                            if agent_id:
                                self.save_agent_contract_stat(
                                    agent_id=agent_id,
                                    client_user_id=client_user_id, 
                                    contract_client_id=client_id,
                                    contract_type=contract_type,
                                    is_first_contract=is_first
                                )
                        except Exception as e:
                            print(f"Ошибка учета статистики агента: {e}")
                        print(f"Новый клиент сохранен в базу с client_id: {client_id}, db_id: {db_id}")
                        
                        # 4. Возвращаем client_id и обновленный data
                        return client_id, data
            except psycopg2.IntegrityError as e:
                if "client_id" in str(e) and attempt < max_retries - 1:
                    print(f"IntegrityError на попытке {attempt + 1}: {e}")
                    time.sleep(0.1 + random.uniform(0, 0.1))
                    continue
                else:
                    raise e
            except Exception as e:
                import traceback
                print(f"Ошибка сохранения клиента на попытке {attempt + 1}")
                print(f"Тип ошибки: {type(e).__name__}")
                print(f"Текст ошибки: {str(e)}")
                print(f"Полный traceback:")
                traceback.print_exc()
                if attempt < max_retries - 1:
                    time.sleep(0.1)
                    continue
                else:
                    raise e
        
        raise Exception("Исчерпаны все попытки сохранения клиента")
    def _determine_contract_stats(self, client_user_id, creator_user_id):
        """Определение агента и типа договора"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    # Проверяем, есть ли связь клиент-агент
                    cursor.execute('''
                        SELECT agent_id FROM client_agent_relationships 
                        WHERE client_user_id = %s
                    ''', (client_user_id,))
                    relationship = cursor.fetchone()
                    
                    if relationship:
                        agent_id = str(relationship['agent_id'])
                        
                        # Если агент сам создает договор - direct
                        if creator_user_id == agent_id:
                            return agent_id, 'direct', False
                        
                        # Если клиент сам создает договор - НЕ начисляем агенту
                        if creator_user_id == client_user_id:
                            return None, None, False
                        
                        # Если клиент создает после приглашения - invited
                        is_first = self.check_if_first_contract_for_invited_client(client_user_id)
                        return agent_id, 'invited', is_first
                    
                    # Если создатель - агент без приглашения
                    cursor.execute('''
                        SELECT user_id FROM admins 
                        WHERE user_id = %s AND admin_value IN ('Агент', 'Клиент', 'Администратор')
                    ''', (creator_user_id,))
                    if cursor.fetchone():
                        return creator_user_id, 'direct', False
                    
                    return None, None, False
                    
        except Exception as e:
            print(f"Ошибка определения типа договора: {e}")
            return None, None, False
    def _prepare_client_data(self, data, created_at):
        """Подготавливает данные клиента для базы данных"""
        
    
        result = {
                'accident': data.get('accident',''),
                'client_id': data.get('client_id',''),
                'fio': data.get('fio', ''),
                'seria_pasport': data.get('seria_pasport', ''),
                'number_pasport': data.get('number_pasport', ''),
                'where_pasport': data.get('where_pasport', ''),
                'when_pasport': data.get('when_pasport', ''),
                'address': data.get('address', ''),
                'index_postal': data.get('index_postal', ''),
                'number': data.get('number', ''),
                'date_of_birth': data.get('date_of_birth', ''),
                'city_birth': data.get('city_birth', ''),
                'date_dtp': data.get('date_dtp', ''),
                'time_dtp': data.get('time_dtp', ''),
                'address_dtp': data.get('address_dtp', ''),
                'who_dtp': data.get('who_dtp', ''),
                'marks': data.get('marks', ''),
                'car_number': data.get('car_number', ''),
                'year_auto': data.get('year_auto', ''),
                'docs': data.get('docs', ''),
                'dkp': data.get('dkp', ''),
                'seria_docs': data.get('seria_docs', ''),
                'number_docs': data.get('number_docs', ''),
                'data_docs': data.get('data_docs', ''),
                'insurance': data.get('insurance', ''),
                'seria_insurance': data.get('seria_insurance', ''),
                'number_insurance': data.get('number_insurance', ''),
                'date_insurance': data.get('date_insurance', ''),
                'fio_culp': data.get('fio_culp', ''),
                'marks_culp': data.get('marks_culp', ''),
                'number_auto_culp': data.get('number_auto_culp', ''),
                'number_photo': data.get('number_photo', ''),
                'place': data.get('place', ''),
                'bank': data.get('bank', ''),
                'bank_account': data.get('bank_account', ''),
                'bank_account_corr': data.get('bank_account_corr', ''),
                'BIK': data.get('BIK', ''),
                'INN': data.get('INN', ''),
                'created_at': created_at,
                'data_json': json.dumps(data, ensure_ascii=False),
                'sobstvenik': data.get('sobstvenik', ''),
                'fio_sobs': data.get('fio_sobs', ''),
                'date_of_birth_sobs': data.get('date_of_birth_sobs',''),
                'answer_ins': data.get('answer_ins', ''),
                'analis_ins': data.get('analis_ins', ''),
                'vibor': data.get('vibor', ''),
                'vibor1': data.get('vibor1', ''),
                'Nv_ins': data.get('Nv_ins', ''),
                'date_coin_ins': data.get('date_coin_ins', ''),
                'Na_ins': data.get('Na_ins', ''),
                'date_Na_ins': data.get('date_Na_ins', ''),
                'date_exp': data.get('date_exp', ''),
                'org_exp': data.get('org_exp', ''),
                'coin_exp': data.get('coin_exp', ''),
                'date_sto': data.get('date_sto', ''),
                'time_sto': data.get('time_sto', ''),
                'address_sto': data.get('address_sto', ''), 
                'coin_exp_izn': data.get('coin_exp_izn', ''), 
                'coin_osago': data.get('coin_osago', ''), 
                'coin_not': data.get('coin_not', ''), 
                'N_dov_not': data.get('N_dov_not', ''), 
                'data_dov_not': data.get('data_dov_not', ''), 
                'fio_not': data.get('fio_not', ''), 
                'number_not': data.get('number_not', ''),
                'date_ins': data.get('date_ins', ''),
                'date_pret': data.get('date_pret', ''),
                'pret': data.get('pret', ''),
                'ombuc': data.get('ombuc', ''),
                'data_pret_prin': data.get('data_pret_prin', ''),
                'data_pret_otv': data.get('data_pret_otv', ''),
                'N_pret_prin': data.get('N_pret_prin', ''),
                'date_ombuc': data.get('date_ombuc', ''),
                'date_ins_pod': data.get('date_ins_pod', ''),
                'seria_vu_culp': data.get('seria_vu_culp', ''),
                'number_vu_culp': data.get('number_vu_culp', ''),
                'data_vu_culp': data.get('data_vu_culp', ''),
                'date_of_birth_culp': data.get('date_of_birth_culp', ''),
                'index_culp': data.get('index_culp', ''),
                'address_culp': data.get('address_culp', ''),
                'number_culp': data.get('number_culp', ''),
                'N_viplat_work': data.get('N_viplat_work', ''),
                'date_viplat_work': data.get('date_viplat_work', ''),
                'N_plat_por': data.get('N_plat_por', ''),
                'date_plat_por': data.get('date_plat_por', ''),
                'sud': data.get('sud', ''),
                'gos_money': data.get('gos_money', ''),
                'date_izvesh_dtp': data.get('date_izvesh_dtp', ''),
                'date_isk': data.get('date_isk', ''),
                'dop_osm': data.get('dop_osm', ''),
                'ev': data.get('ev', ''),
                'address_park': data.get('address_park', ''),
                'fio_k':data.get('fio_k', ''),
                'data_dop_osm': data.get('data_dop_osm', ''),
                'viborRem': data.get('viborRem', ''),
                'date_zayav_sto':data.get('date_zayav_sto', ''),
                'pret_sto':data.get('pret_sto', ''),
                'data_otkaz_sto':data.get('data_otkaz_sto', ''),
                'date_napr_sto': data.get('date_napr_sto', ''),
                'address_sto_main': data.get('address_sto_main', ''),
                'data_sto_main': data.get('data_sto_main', ''),
                'time_sto_main': data.get('time_sto_main', ''),
                'city_sto': data.get('city_sto', ''),
                'Done': data.get('Done', ''),
                'city': data.get('city', ''),
                'year': data.get('year', ''),
                'street': data.get('street', ''),
                'N_gui': data.get('N_gui', ''),
                'date_gui': data.get('date_gui', ''),
                'N_prot': data.get('N_prot', ''),
                'date_prot': data.get('date_prot', ''),
                'date_road': data.get('date_road', ''),
                'N_kv_not': data.get('N_kv_not', ''),
                'date_kv_not': data.get('date_kv_not', ''),
                'N_kv_ur': data.get('N_kv_ur', ''),
                'date_kv_ur': data.get('date_kv_ur', ''),
                'N_kv_exp': data.get('N_kv_exp', ''),
                'status': data.get('status', ''),
                'fio_c': data.get('fio_c', ''),
                'fio_c_k': data.get('fio_c_k', ''),
                'seria_pasport_c': data.get('seria_pasport_c', ''),
                'number_pasport_c': data.get('number_pasport_c', ''),
                'where_pasport_c': data.get('where_pasport_c', ''),
                'when_pasport_c': data.get('when_pasport_c', ''),
                'address_c': data.get('address_c', ''),
                'date_of_birth_c': data.get('date_of_birth_c', ''),
                'coin_c': data.get('coin_c', ''),
                'city_birth_c': data.get('city_birth_c', ''),
                'index_postal_c': data.get('index_postal_c', ''),
                'number_c': data.get('number_c', ''),
                'money_exp': data.get('money_exp', ''),
                'user_id': data.get('user_id', ''),
                'agent_id': data.get('agent_id', ''),
                'ur_money': data.get('ur_money', ''),
                'calculation': data.get('calculation', ''),
        }
        print(f"DEBUG: Количество полей в _prepare_client_data: {len(result)}")
        return result
    def get_pending_approvals_count(self, document_type=None):
        """Получить количество документов на подтверждение"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    if document_type:
                        cursor.execute("""
                            SELECT COUNT(*) FROM pending_approvals 
                            WHERE status = 'pending' AND document_type = %s
                        """, (document_type,))
                    else:
                        cursor.execute("""
                            SELECT COUNT(*) FROM pending_approvals 
                            WHERE status = 'pending'
                        """)
                    return cursor.fetchone()[0]
        except Exception as e:
            print(f"Ошибка получения количества документов: {e}")
            return 0

    def get_pending_approvals_list(self, document_type):
        """Получить список документов на подтверждение"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    cursor.execute("""
                        SELECT * FROM pending_approvals 
                        WHERE status = 'pending' AND document_type = %s
                        ORDER BY created_at DESC
                    """, (document_type,))
                    return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            print(f"Ошибка получения списка документов: {e}")
            return []

    def update_approval_status(self, approval_id, status, reviewed_by):
        """Обновить статус документа"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute("""
                        UPDATE pending_approvals 
                        SET status = %s, reviewed_by = %s, reviewed_at = CURRENT_TIMESTAMP
                        WHERE id = %s
                    """, (status, reviewed_by, approval_id))
                    conn.commit()
                    return True
        except Exception as e:
            print(f"Ошибка обновления статуса: {e}")
            return False

    def add_agent_earning(self, agent_id, amount):
        """Начислить заработок агенту"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    # Создаем запись если не существует
                    cursor.execute("""
                        INSERT INTO agent_finances (agent_id, balance, total_earned)
                        VALUES (%s, %s, %s)
                        ON CONFLICT (agent_id) DO UPDATE
                        SET balance = agent_finances.balance + %s,
                            total_earned = agent_finances.total_earned + %s,
                            last_updated = CURRENT_TIMESTAMP
                    """, (agent_id, amount, amount, amount, amount))
                    conn.commit()
                    return True
        except Exception as e:
            print(f"Ошибка начисления заработка: {e}")
            return False

    def get_agent_balance(self, agent_id):
        """Получить баланс агента"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    cursor.execute("""
                        SELECT balance, total_earned FROM agent_finances 
                        WHERE agent_id = %s
                    """, (agent_id,))
                    result = cursor.fetchone()
                    if result:
                        return dict(result)
                    else:
                        # Создаем запись с нулевым балансом
                        cursor.execute("""
                            INSERT INTO agent_finances (agent_id, balance, total_earned)
                            VALUES (%s, 0, 0)
                            RETURNING balance, total_earned
                        """, (agent_id,))
                        conn.commit()
                        return {'balance': 0, 'total_earned': 0}
        except Exception as e:
            print(f"Ошибка получения баланса: {e}")
            return {'balance': 0, 'total_earned': 0}

    def create_withdrawal_request(self, agent_id, agent_fio, amount):
        """Создать заявку на вывод средств"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute("""
                        INSERT INTO withdrawal_requests (agent_id, agent_fio, amount)
                        VALUES (%s, %s, %s)
                        RETURNING id
                    """, (agent_id, agent_fio, amount))
                    conn.commit()
                    return cursor.fetchone()[0]
        except Exception as e:
            print(f"Ошибка создания заявки на вывод: {e}")
            return None

    def get_pending_withdrawals(self):
        """Получить список заявок на вывод"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    cursor.execute("""
                        SELECT * FROM withdrawal_requests 
                        WHERE status = 'pending'
                        ORDER BY created_at DESC
                    """)
                    return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            print(f"Ошибка получения заявок на вывод: {e}")
            return []

    def process_withdrawal(self, withdrawal_id, status, reviewed_by, agent_id, amount):
        """Обработать заявку на вывод"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute("""
                        UPDATE withdrawal_requests 
                        SET status = %s, reviewed_by = %s, reviewed_at = CURRENT_TIMESTAMP
                        WHERE id = %s
                    """, (status, reviewed_by, withdrawal_id))
                    
                    if status == 'approved':
                        # Вычитаем из баланса агента
                        cursor.execute("""
                            UPDATE agent_finances 
                            SET balance = balance - %s,
                                last_updated = CURRENT_TIMESTAMP
                            WHERE agent_id = %s
                        """, (amount, agent_id))
                    
                    conn.commit()
                    return True
        except Exception as e:
            print(f"Ошибка обработки заявки на вывод: {e}")
            return False

    def get_agent_monthly_earning(self, agent_id):
        """Получить заработок агента за текущий месяц по начислениям"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute("""
                        SELECT COALESCE(SUM(amount), 0) as monthly_earning
                        FROM agent_earnings_history 
                        WHERE agent_id = %s 
                        AND EXTRACT(MONTH FROM payment_confirmed_at) = EXTRACT(MONTH FROM CURRENT_DATE)
                        AND EXTRACT(YEAR FROM payment_confirmed_at) = EXTRACT(YEAR FROM CURRENT_DATE)
                    """, (agent_id,))
                    result = cursor.fetchone()
                    return float(result[0]) if result else 0.0
        except Exception as e:
            print(f"Ошибка получения месячного заработка: {e}")
            return 0.0
    def update_client_contract_relationship(self, client_user_id, client_contract_id):
        """Обновление связи клиент-агент с добавлением contract_id после подписания договора"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(
                        """UPDATE client_agent_relationships 
                        SET client_contract_id = %s 
                        WHERE client_user_id = %s""",
                        (client_contract_id, client_user_id)
                    )
                    conn.commit()
                    print(f"Обновлена связь для клиента {client_user_id}: contract_id = {client_contract_id}")
                    return True
        except Exception as e:
            print(f"Ошибка обновления связи клиент-договор: {e}")
            return False
    def get_client_by_client_id(self, client_id):
        """Получение клиента по client_id"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    cursor.execute("SELECT * FROM clients WHERE client_id = %s", (client_id,))
                    result = cursor.fetchone()
                    
                    if result:
                        return dict(result)
                    return None
        except Exception as e:
            print(f"Ошибка получения клиента: {e}")
            return None
    def get_clients_by_agent_id(self, agent_id):
        """Получить всех клиентов конкретного агента"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    cursor.execute("""
                        SELECT c.*, car.created_at as invite_date
                        FROM clients c 
                        JOIN client_agent_relationships car ON c.user_id::text = car.client_user_id::text
                        WHERE car.agent_id = %s
                        ORDER BY c.created_at DESC
                    """, (agent_id,))
                    
                    return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            print(f"Ошибка получения клиентов агента: {e}")
            return []
    def search_clients(self, search_term):
        """Улучшенный поиск клиентов по ФИО, номеру телефона, номеру авто или client_id"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    search_term = search_term.strip()
                    results = []
                    
                    # 1. Поиск точного совпадения по всем полям
                    exact_patterns = [
                        search_term,
                        search_term.lower(),
                        search_term.upper(),
                        search_term.title()
                    ]
                    
                    for pattern in exact_patterns:
                        search_query = '''
                        SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                               COALESCE(data_json, '{}') as data_json
                        FROM clients 
                        WHERE fio = %s OR number = %s OR car_number = %s OR client_id = %s
                        ORDER BY created_at DESC
                        '''
                        
                        cursor.execute(search_query, (pattern, pattern, pattern, pattern))
                        exact_results = cursor.fetchall()
                        if exact_results:
                            results.extend(exact_results)
                            break
                    
                    # 2. Поиск частичного совпадения
                    if not results:
                        partial_patterns = [
                            f"%{search_term}%",
                            f"%{search_term.lower()}%",
                            f"%{search_term.upper()}%",
                            f"%{search_term.title()}%"
                        ]
                        
                        for pattern in partial_patterns:
                            search_query = '''
                            SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                                   COALESCE(data_json, '{}') as data_json
                            FROM clients 
                            WHERE fio ILIKE %s OR number ILIKE %s OR car_number ILIKE %s OR client_id ILIKE %s
                            ORDER BY created_at DESC
                            '''
                            
                            cursor.execute(search_query, (pattern, pattern, pattern, pattern))
                            partial_results = cursor.fetchall()
                            if partial_results:
                                results.extend(partial_results)
                                break
                    
                    # 3. Поиск по отдельным словам (для ФИО)
                    if not results:
                        search_words = search_term.split()
                        if len(search_words) >= 2:
                            first_word = search_words[0].strip()
                            second_word = search_words[1].strip()
                            
                            # Различные варианты регистра
                            word_variants = []
                            for word in [first_word, second_word]:
                                word_variants.append([
                                    word,
                                    word.lower(),
                                    word.upper(),
                                    word.title()
                                ])
                            
                            # Пробуем все комбинации
                            for first_variants in word_variants[0]:
                                for second_variants in word_variants[1]:
                                    search_query = '''
                                    SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                                           COALESCE(data_json, '{}') as data_json
                                    FROM clients 
                                    WHERE fio ILIKE %s AND fio ILIKE %s
                                    ORDER BY created_at DESC
                                    '''
                                    
                                    cursor.execute(search_query, (f"%{first_variants}%", f"%{second_variants}%"))
                                    word_results = cursor.fetchall()
                                    if word_results:
                                        results.extend(word_results)
                                        break
                                
                                if results:
                                    break
                    
                    # 4. Полнотекстовый поиск по русскому языку для ФИО
                    if not results:
                        search_query = '''
                        SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                               COALESCE(data_json, '{}') as data_json
                        FROM clients 
                        WHERE to_tsvector('russian', fio) @@ plainto_tsquery('russian', %s)
                        ORDER BY ts_rank(to_tsvector('russian', fio), plainto_tsquery('russian', %s)) DESC
                        '''
                        
                        cursor.execute(search_query, (search_term, search_term))
                        fulltext_results = cursor.fetchall()
                        if fulltext_results:
                            results.extend(fulltext_results)
                    
                    # Удаляем дубликаты по client_id
                    unique_results = []
                    seen_client_ids = set()
                    
                    for result in results:
                        client_id = result['client_id']
                        if client_id not in seen_client_ids:
                            unique_results.append(dict(result))
                            seen_client_ids.add(client_id)
                    
                    return unique_results
        except Exception as e:
            print(f"Ошибка поиска клиентов: {e}")
            return []
    
    def search_clients_by_fio(self, search_term):
        """Специализированный поиск только по ФИО с расширенными возможностями"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    search_term = search_term.strip()
                    results = []
                    
                    print(f"Поиск по ФИО: '{search_term}'")
                    
                    # Функция для замены ё на е и наоборот
                    def get_e_yo_variants(text):
                        variants = set()
                        variants.add(text)  # оригинал
                        
                        # Замена ё на е
                        if 'ё' in text.lower():
                            variants.add(text.replace('ё', 'е').replace('Ё', 'Е'))
                        
                        # Замена е на ё (только в позициях где может быть ё)
                        if 'е' in text.lower():
                            # Простая замена всех е на ё
                            variants.add(text.replace('е', 'ё').replace('Е', 'Ё'))
                        
                        return list(variants)
                    
                    # Генерируем варианты с учетом ё/е для поискового термина
                    search_variants = get_e_yo_variants(search_term)
                    print(f"Варианты поиска с ё/е: {search_variants}")
                    
                    # 1. Точное совпадение (с учетом ё/е)
                    exact_patterns = set()
                    for variant in search_variants:
                        exact_patterns.add(variant)
                        exact_patterns.add(variant.lower())
                        exact_patterns.add(variant.upper())
                        exact_patterns.add(variant.title())
                    
                    for pattern in exact_patterns:
                        query = '''
                        SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                            COALESCE(data_json, '{}') as data_json
                        FROM clients 
                        WHERE fio = %s
                        ORDER BY id DESC
                        '''
                        
                        cursor.execute(query, (pattern,))
                        exact_results = cursor.fetchall()
                        if exact_results:
                            results.extend(exact_results)
                            print(f"Найдено точных совпадений для '{pattern}': {len(exact_results)}")
                    
                    # 2. Частичное совпадение (с учетом ё/е)
                    if not results:
                        partial_patterns = set()
                        for variant in search_variants:
                            partial_patterns.add(f"%{variant}%")
                            partial_patterns.add(f"%{variant.lower()}%")
                            partial_patterns.add(f"%{variant.upper()}%")
                            partial_patterns.add(f"%{variant.title()}%")
                        
                        for pattern in partial_patterns:
                            query = '''
                            SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                                COALESCE(data_json, '{}') as data_json
                            FROM clients 
                            WHERE fio ILIKE %s
                            ORDER BY id DESC
                            '''
                            
                            cursor.execute(query, (pattern,))
                            partial_results = cursor.fetchall()
                            if partial_results:
                                results.extend(partial_results)
                                print(f"Найдено частичных совпадений для '{pattern}': {len(partial_results)}")
                    
                    # 3. Поиск по отдельным словам (с учетом ё/е)
                    if not results:
                        search_words = search_term.split()
                        if len(search_words) >= 2:
                            first_word = search_words[0].strip()
                            second_word = search_words[1].strip()
                            
                            # Варианты с ё/е для каждого слова
                            first_word_variants = get_e_yo_variants(first_word)
                            second_word_variants = get_e_yo_variants(second_word)
                            
                            # Пробуем все комбинации
                            for first_variant in first_word_variants:
                                for second_variant in second_word_variants:
                                    # Различные варианты регистра
                                    for first_case in [first_variant, first_variant.lower(), first_variant.upper(), first_variant.title()]:
                                        for second_case in [second_variant, second_variant.lower(), second_variant.upper(), second_variant.title()]:
                                            query = '''
                                            SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                                                COALESCE(data_json, '{}') as data_json
                                            FROM clients 
                                            WHERE fio ILIKE %s AND fio ILIKE %s
                                            ORDER BY id DESC
                                            '''
                                            
                                            cursor.execute(query, (f"%{first_case}%", f"%{second_case}%"))
                                            word_results = cursor.fetchall()
                                            if word_results:
                                                results.extend(word_results)
                                                print(f"Найдено по словам '{first_case}' + '{second_case}': {len(word_results)}")
                                                break
                                
                                if results:
                                    break
                    
                    # 4. Поиск только по первому слову (фамилии) с учетом ё/е
                    if not results:
                        first_word = search_term.split()[0] if search_term.split() else search_term
                        first_word_variants = get_e_yo_variants(first_word)
                        
                        for variant in first_word_variants:
                            # Различные варианты регистра
                            for case_variant in [variant, variant.lower(), variant.upper(), variant.title()]:
                                query = '''
                                SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                                    COALESCE(data_json, '{}') as data_json
                                FROM clients 
                                WHERE fio ILIKE %s
                                ORDER BY id DESC
                                '''
                                
                                cursor.execute(query, (f"%{case_variant}%",))
                                surname_results = cursor.fetchall()
                                if surname_results:
                                    results.extend(surname_results)
                                    print(f"Найдено по фамилии '{case_variant}': {len(surname_results)}")
                                    break
                    
                    # 5. Полнотекстовый поиск по русскому языку (PostgreSQL сам обрабатывает ё/е)
                    if not results:
                        query = '''
                        SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                            COALESCE(data_json, '{}') as data_json
                        FROM clients 
                        WHERE to_tsvector('russian', fio) @@ plainto_tsquery('russian', %s)
                        ORDER BY ts_rank(to_tsvector('russian', fio), plainto_tsquery('russian', %s)) DESC
                        '''
                        
                        cursor.execute(query, (search_term, search_term))
                        fulltext_results = cursor.fetchall()
                        if fulltext_results:
                            results.extend(fulltext_results)
                            print(f"Найдено полнотекстовым поиском: {len(fulltext_results)}")
                    
                    # Удаляем дубликаты по client_id
                    unique_results = []
                    seen_client_ids = set()
                    
                    for result in results:
                        client_id = result['client_id']
                        if client_id not in seen_client_ids:
                            unique_results.append(dict(result))
                            seen_client_ids.add(client_id)
                    
                    print(f"Уникальных результатов поиска по ФИО: {len(unique_results)}")
                    
                    return unique_results
        except Exception as e:
            print(f"Ошибка поиска по ФИО: {e}")
            return []
    def get_client_by_user_id(self, user_id):
        """Получение клиента по user_id"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    cursor.execute("SELECT * FROM clients WHERE user_id = %s::text ORDER BY created_at DESC LIMIT 1", (user_id,))
                    result = cursor.fetchone()
                    return dict(result) if result else None
        except Exception as e:
            print(f"Ошибка получения клиента по user_id: {e}")
            return None

    def get_agent_statistics(self, user_id):
        """Статистика клиентов агента"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute("""
                        SELECT 
                            COUNT(CASE WHEN "Done" != 'Yes' OR "Done" IS NULL THEN 1 END) as in_progress,
                            COUNT(CASE WHEN "Done" = 'Yes' THEN 1 END) as completed,
                            COUNT(*) as total
                        FROM clients WHERE user_id = %s::text
                    """, (user_id,))
                    result = cursor.fetchone()
                    return {
                        'in_progress': result[0] or 0,
                        'completed': result[1] or 0,
                        'total': result[2] or 0
                    }
        except Exception as e:
            print(f"Ошибка получения статистики агента: {e}")
            return {'in_progress': 0, 'completed': 0, 'total': 0}

    def get_city_statistics(self, city):
        """Статистика клиентов по городу"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute("""
                        SELECT 
                            COUNT(CASE WHEN "Done" != 'Yes' OR "Done" IS NULL THEN 1 END) as in_progress,
                            COUNT(CASE WHEN "Done" = 'Yes' THEN 1 END) as completed,
                            COUNT(*) as total
                        FROM clients WHERE city = %s
                    """, (city,))
                    result = cursor.fetchone()
                    return {
                        'in_progress': result[0] or 0,
                        'completed': result[1] or 0,
                        'total': result[2] or 0
                    }
        except Exception as e:
            print(f"Ошибка получения статистики по городу: {e}")
            return {'in_progress': 0, 'completed': 0, 'total': 0}

    def get_global_statistics(self):
        """Общая статистика по всем клиентам"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute("""
                        SELECT 
                            COUNT(CASE WHEN "Done" != 'Yes' OR "Done" IS NULL THEN 1 END) as in_progress,
                            COUNT(CASE WHEN "Done" = 'Yes' THEN 1 END) as completed,
                            COUNT(*) as total
                        FROM clients
                    """)
                    result = cursor.fetchone()
                    return {
                        'in_progress': result[0] or 0,
                        'completed': result[1] or 0,
                        'total': result[2] or 0
                    }
        except Exception as e:
            print(f"Ошибка получения общей статистики: {e}")
            return {'in_progress': 0, 'completed': 0, 'total': 0}
    def get_database_stats(self):
        """Статистика базы данных"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    # Общее количество клиентов
                    cursor.execute("SELECT COUNT(*) FROM clients")
                    total_clients = cursor.fetchone()[0]
                    
                    # Последний client_id
                    cursor.execute('''
                    SELECT client_id FROM clients 
                    WHERE client_id LIKE '70%' 
                    ORDER BY CAST(client_id AS INTEGER) DESC 
                    LIMIT 1
                    ''')
                    result = cursor.fetchone()
                    last_client_id = result[0] if result else "Нет клиентов"
                    
                    # Размер базы данных
                    cursor.execute('''
                    SELECT pg_size_pretty(pg_database_size(current_database())) AS size
                    ''')
                    db_size = cursor.fetchone()[0]
                    
                    # Размер таблицы clients
                    cursor.execute('''
                    SELECT pg_size_pretty(pg_total_relation_size('clients')) AS size
                    ''')
                    table_size = cursor.fetchone()[0]
                    
                    return {
                        'total_clients': total_clients,
                        'last_client_id': last_client_id,
                        'db_size': db_size,
                        'table_size': table_size
                    }
        except Exception as e:
            print(f"Ошибка получения статистики: {e}")
            return {
                'total_clients': 0,
                'last_client_id': "Ошибка",
                'db_size': "Ошибка",
                'table_size': "Ошибка"
            }
    def check_admin_exists(self, user_id):
        """Проверка существования администратора"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute("SELECT COUNT(*) FROM admins WHERE user_id = %s::text AND is_active = true", (user_id,))
                    return cursor.fetchone()[0] > 0
        except Exception as e:
            print(f"Ошибка проверки администратора: {e}")
            return False

    def save_admin(self, admin_data):
        """Сохранение данных администратора"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    created_at = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
                    
                    insert_query = '''
                    INSERT INTO admins (
                        user_id, fio, fio_k, seria_pasport, number_pasport, 
                        where_pasport, when_pasport, date_of_birth, city_birth,
                        address, index_postal, admin_value, city_admin, number, created_at,
                        invited_by_user_id, invited_by_type
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (user_id) DO UPDATE SET
                        fio = EXCLUDED.fio,
                        fio_k = EXCLUDED.fio_k,
                        seria_pasport = EXCLUDED.seria_pasport,
                        number_pasport = EXCLUDED.number_pasport,
                        where_pasport = EXCLUDED.where_pasport,
                        when_pasport = EXCLUDED.when_pasport,
                        date_of_birth = EXCLUDED.date_of_birth,
                        city_birth = EXCLUDED.city_birth,
                        address = EXCLUDED.address,
                        index_postal = EXCLUDED.index_postal,
                        admin_value = EXCLUDED.admin_value,
                        city_admin = EXCLUDED.city_admin,
                        number = EXCLUDED.number,
                        invited_by_user_id = EXCLUDED.invited_by_user_id,
                        invited_by_type = EXCLUDED.invited_by_type
                    '''
                    
                    cursor.execute(insert_query, (
                        admin_data.get('user_id'),
                        admin_data.get('fio'),
                        admin_data.get('fio_k', ''),
                        admin_data.get('seria_pasport'),
                        admin_data.get('number_pasport'),
                        admin_data.get('where_pasport'),
                        admin_data.get('when_pasport'),
                        admin_data.get('date_of_birth', ''),
                        admin_data.get('city_birth', ''),
                        admin_data.get('address', ''),
                        admin_data.get('index_postal', ''),
                        admin_data.get('admin_value'),
                        admin_data.get('city_admin'),
                        admin_data.get('number', ''),
                        created_at,
                        admin_data.get('invited_by_user_id', None),
                        admin_data.get('invited_by_type', None)
                    ))
                    
                    conn.commit()
                    print(f"✅ Администратор {admin_data.get('fio')} успешно сохранен в БД")
                    return True
        except Exception as e:
            print(f"❌ Ошибка сохранения администратора: {e}")
            import traceback
            traceback.print_exc()
            return False
    def update_admin(self, admin_data):
        """Обновление данных существующего администратора"""
        try:
            with self.get_connection() as conn:
                with conn.cursor() as cursor:
                    
                    update_query = '''
                    UPDATE admins SET
                        fio = %s,
                        fio_k = %s,
                        seria_pasport = %s,
                        number_pasport = %s,
                        where_pasport = %s,
                        when_pasport = %s,
                        date_of_birth = %s,
                        city_birth = %s,
                        address = %s,
                        index_postal = %s,
                        admin_value = %s,
                        city_admin = %s,
                        number = %s,
                        invited_by_user_id = %s,
                        invited_by_type = %s
                    WHERE user_id = %s
                    '''
                    
                    cursor.execute(update_query, (
                        admin_data.get('fio'),
                        admin_data.get('fio_k', ''),
                        admin_data.get('seria_pasport'),
                        admin_data.get('number_pasport'),
                        admin_data.get('where_pasport'),
                        admin_data.get('when_pasport'),
                        admin_data.get('date_of_birth', ''),
                        admin_data.get('city_birth', ''),
                        admin_data.get('address', ''),
                        admin_data.get('index_postal', ''),
                        admin_data.get('admin_value'),
                        admin_data.get('city_admin'),
                        admin_data.get('number', ''),
                        admin_data.get('invited_by_user_id', None),
                        admin_data.get('invited_by_type', None),
                        admin_data.get('user_id')  # WHERE условие
                    ))
                    
                    # Проверяем, была ли обновлена хотя бы одна строка
                    if cursor.rowcount > 0:
                        conn.commit()
                        print(f"✅ Данные администратора {admin_data.get('fio')} успешно обновлены")
                        return True
                    else:
                        print(f"⚠️ Администратор с user_id {admin_data.get('user_id')} не найден")
                        return False
                        
        except Exception as e:
            print(f"❌ Ошибка обновления администратора: {e}")
            import traceback
            traceback.print_exc()
            return False
    def get_admin_by_user_id(self, user_id):
        """Получение данных администратора по user_id"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    cursor.execute("""
                        SELECT id, user_id, fio, fio_k, seria_pasport, number_pasport, 
                            where_pasport, when_pasport, date_of_birth, city_birth,
                            address, index_postal, admin_value, city_admin, number,
                            created_at, is_active, invited_by_user_id, invited_by_type
                        FROM admins 
                        WHERE user_id = %s::text AND is_active = true
                    """, (user_id,))
                    
                    result = cursor.fetchone()
                    
                    if result:
                        return dict(result)
                    return None
        except Exception as e:
            print(f"Ошибка получения данных администратора: {e}")
            return None
    def get_admin_by_fio(self, fio):
        """Получение данных администратора по user_id"""
        try:
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    cursor.execute("""
                        SELECT id, user_id, fio, fio_k, seria_pasport, number_pasport, 
                            where_pasport, when_pasport, date_of_birth, city_birth,
                            address, index_postal, admin_value, city_admin, number,
                            created_at, is_active, invited_by_user_id, invited_by_type
                        FROM admins 
                        WHERE fio = %s::text AND is_active = true
                    """, (fio,))
                    
                    result = cursor.fetchone()
                    
                    if result:
                        return dict(result)
                    return None
        except Exception as e:
            print(f"Ошибка получения данных администратора: {e}")
            return None
    def export_admins_to_excel(self, file_path, city_filter=None):
        """Экспорт данных администраторов в Excel файл с количеством клиентов"""
        try:
            import pandas as pd
            
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    if city_filter:
                        query = """
                        SELECT a.user_id, a.fio, a.seria_pasport, a.number_pasport, 
                            a.where_pasport, a.when_pasport, a.admin_value, a.city_admin, 
                            a.created_at, a.is_active, a.date_of_birth, a.number,
                            COUNT(c.client_id) as client_count
                        FROM admins a
                        LEFT JOIN clients c ON a.user_id::text = c.user_id
                        WHERE a.city_admin = %s AND a.is_active = true AND a.admin_value != 'Клиент'
                        GROUP BY a.user_id, a.fio, a.seria_pasport, a.number_pasport, 
                                a.where_pasport, a.when_pasport, a.admin_value, a.city_admin, 
                                a.created_at, a.is_active, a.date_of_birth, a.number
                        ORDER BY a.created_at DESC
                        """
                        cursor.execute(query, (city_filter,))
                    else:
                        query = """
                        SELECT a.user_id, a.fio, a.seria_pasport, a.number_pasport, 
                            a.where_pasport, a.when_pasport, a.admin_value, a.city_admin, 
                            a.created_at, a.is_active, a.date_of_birth, a.number
                            COUNT(c.client_id) as client_count
                        FROM admins a
                        LEFT JOIN clients c ON a.user_id::text = c.user_id
                        WHERE a.is_active = true AND a.admin_value != 'Клиент'
                        GROUP BY a.user_id, a.fio, a.seria_pasport, a.number_pasport, 
                                a.where_pasport, a.when_pasport, a.admin_value, a.city_admin, 
                                a.created_at, a.is_active, a.date_of_birth, a.number
                        ORDER BY a.created_at DESC
                        """
                        cursor.execute(query)
                    
                    admins = cursor.fetchall()
                    
                    if not admins:
                        print("Нет данных администраторов для экспорта")
                        return False
                    
                    # Преобразуем в DataFrame
                    df = pd.DataFrame([dict(admin) for admin in admins])
                    
                    # Переименовываем колонки для читаемости
                    column_mapping = {
                        'admin_value': 'Должность',
                        'fio': 'ФИО',
                        'date_of_birth': 'Дата рождения',
                        'city_admin': 'Город',
                        'number': 'Номер телефона',
                        'client_count': 'Количество обработанных договоров'
                    }
                    
                    df = df.rename(columns=column_mapping)
                    
                    # Экспортируем в Excel
                    df.to_excel(file_path, index=False, engine='openpyxl')
                    
                    workbook = load_workbook(file_path)
                    worksheet = workbook.active
                    
                    # Автоподбор ширины для каждого столбца
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        
                        for cell in column:
                            try:
                                # Получаем длину текста в ячейке
                                if cell.value:
                                    cell_length = len(str(cell.value))
                                    if cell_length > max_length:
                                        max_length = cell_length
                            except:
                                pass
                        
                        # Устанавливаем ширину столбца с запасом
                        adjusted_width = min(max_length + 2, 50)  # Максимальная ширина 50
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Сохраняем изменения
                    workbook.save(file_path)
                    
                    print(f"Экспорт {len(admins)} администраторов завершен: {file_path}")
                    return True
                    
        except Exception as e:
            print(f"Ошибка экспорта администраторов: {e}")
            return False
    
    def export_clients_to_excel_by_city(self, file_path, city_filter):
        """Экспорт клиентов по городу в Excel файл с данными администратора"""
        try:
            import pandas as pd
            
            with self.get_connection() as conn:
                with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                    query = """
                    SELECT c.client_id, c.fio, c.number, c.car_number, c.date_dtp, 
                        c.accident, c.insurance, c.city, c.created_at, c.status,
                        c.seria_pasport, c.number_pasport, c.address, 
                        c.date_of_birth, c.user_id,
                        a.fio as admin_fio
                    FROM clients c
                    LEFT JOIN admins a ON c.user_id = a.user_id::text AND a.is_active = true
                    WHERE c.city = %s
                    ORDER BY c.created_at DESC
                    """
                    cursor.execute(query, (city_filter,))
                    
                    clients = cursor.fetchall()
                    
                    if not clients:
                        print(f"Нет клиентов в городе {city_filter} для экспорта")
                        return False
                    
                    # Преобразуем в DataFrame
                    df = pd.DataFrame([dict(client) for client in clients])
                    
                    # Переименовываем колонки для читаемости
                    column_mapping = {
                        'client_id': 'ID клиента',
                        'fio': 'ФИО клиента',
                        'number': 'Телефон',
                        'car_number': 'Номер авто',
                        'date_dtp': 'Дата ДТП',
                        'accident': 'Тип обращения',
                        'insurance': 'Страховая',
                        'city': 'Город',
                        'created_at': 'Дата создания',
                        'status': 'Статус',
                        'seria_pasport': 'Серия паспорта',
                        'number_pasport': 'Номер паспорта',
                        'address': 'Адрес',
                        'date_of_birth': 'Дата рождения',
                        'user_id': 'ID администратора',
                        'admin_fio': 'ФИО администратора'
                    }
                    
                    df = df.rename(columns=column_mapping)
                    
                    # Экспортируем в Excel
                    df.to_excel(file_path, index=False, engine='openpyxl')
                    
                    print(f"Экспорт {len(clients)} клиентов города {city_filter} завершен: {file_path}")
                    return True
                    
        except Exception as e:
            print(f"Ошибка экспорта клиентов по городу: {e}")
            return False
# Обновленные функции для интеграции с ботом
def save_client_to_db_with_id(data, connection_params=None):
    """Сохранение данных клиента в базу с генерацией client_id"""
    db = DatabaseManager(connection_params)
    return db.save_client_data_with_generated_id(data)
# Создание для интеграции с ботом
def save_client_to_db_with_id_new(data, connection_params=None):
    """Сохранение данных клиента в базу с генерацией client_id"""
    db = DatabaseManager(connection_params)
    return db.save_client_data_with_generated_id_new(data)
def search_clients_in_db(search_term, connection_params=None):
    """Поиск клиентов в базе (универсальный)"""
    db = DatabaseManager(connection_params)
    return db.search_clients(search_term)

def search_clients_by_fio_in_db(search_term, connection_params=None):
    """Поиск клиентов в базе только по ФИО"""
    db = DatabaseManager(connection_params)
    return db.search_clients_by_fio(search_term)

def get_client_from_db_by_client_id(client_id, connection_params=None):
    """Получение клиента по client_id"""
    db = DatabaseManager(connection_params)
    return db.get_client_by_client_id(client_id)

def get_db_stats(connection_params=None):
    """Получение статистики базы данных"""
    db = DatabaseManager(connection_params)
    return db.get_database_stats()
def get_admin_from_db_by_user_id(user_id, connection_params=None):
    """Получение данных администратора по user_id"""
    db = DatabaseManager(connection_params)
    return db.get_admin_by_user_id(user_id)
def get_admin_from_db_by_fio(fio, connection_params=None):
    """Получение данных администратора по user_id"""
    db = DatabaseManager(connection_params)
    return db.get_admin_by_fio(fio)
def export_all_admins_to_excel(file_path, connection_params=None):
    """Экспорт всех администраторов в Excel"""
    db = DatabaseManager(connection_params)
    return db.export_admins_to_excel(file_path)

def export_city_admins_to_excel(file_path, city_filter, connection_params=None):
    """Экспорт администраторов по городу в Excel"""
    db = DatabaseManager(connection_params)
    return db.export_admins_to_excel(file_path, city_filter)
def export_city_clients_to_excel_table(file_path, city_filter, connection_params=None):
    """Экспорт клиентов по городу в Excel"""
    db = DatabaseManager(connection_params)
    return db.export_clients_to_excel_by_city(file_path, city_filter)
def get_client_by_user_id_db(user_id, connection_params=None):
    """Получение клиента по user_id"""
    db = DatabaseManager(connection_params)
    return db.get_client_by_user_id(user_id)

def get_agent_statistics(user_id, connection_params=None):
    """Статистика агента"""
    db = DatabaseManager(connection_params)
    return db.get_agent_statistics(user_id)

def get_city_statistics(city, connection_params=None):
    """Статистика по городу"""
    db = DatabaseManager(connection_params)
    return db.get_city_statistics(city)
def get_invited_clients_count_by_user(user_id, connection_params=None):
    """Получение количества приглашенных клиентов"""
    db = DatabaseManager(connection_params)
    return db.get_invited_clients_count(user_id)

def get_global_statistics(connection_params=None):
    """Общая статистика"""
    db = DatabaseManager(connection_params)
    return db.get_global_statistics()
def get_detailed_agent_statistics(user_id, connection_params=None):
    """Детальная статистика агента"""
    db = DatabaseManager(connection_params)
    return db.get_agent_contract_statistics_detailed(str(user_id))
def get_client_contracts_list(user_id, connection_params=None):
    """Получение списка договоров клиента"""
    db = DatabaseManager(connection_params)
    return db.get_client_contracts(user_id)
def update_client_agent_contract_link(client_user_id, client_contract_id, connection_params=None):
    """Обновление связи клиент-агент с contract_id"""
    db = DatabaseManager(connection_params)
    return db.update_client_contract_relationship(client_user_id, client_contract_id)
# Дополнительные утилиты для миграции из SQLite
class MigrationManager:
    """Класс для миграции данных из SQLite в PostgreSQL"""
    
    def __init__(self, sqlite_path, postgres_params):
        self.sqlite_path = sqlite_path
        self.postgres_params = postgres_params
        self.postgres_db = DatabaseManager(postgres_params)
    
    def migrate_from_sqlite(self):
        """Миграция всех данных из SQLite в PostgreSQL"""
        try:
            import sqlite3
            
            # Подключение к SQLite
            sqlite_conn = sqlite3.connect(self.sqlite_path)
            sqlite_conn.row_factory = sqlite3.Row
            sqlite_cursor = sqlite_conn.cursor()
            
            # Получение всех записей из SQLite
            sqlite_cursor.execute("SELECT * FROM clients ORDER BY id")
            sqlite_records = sqlite_cursor.fetchall()
            
            print(f"Найдено {len(sqlite_records)} записей для миграции")
            
            migrated_count = 0
            error_count = 0
            
            for record in sqlite_records:
                try:
                    # Преобразование записи SQLite в словарь
                    data = dict(record)
                    
                    # Парсинг JSON данных если они есть
                    if data.get('data_json'):
                        try:
                            json_data = json.loads(data['data_json'])
                            # Объединяем с основными данными
                            data.update(json_data)
                        except (json.JSONDecodeError, TypeError):
                            pass
                    
                    # Удаляем id из SQLite, так как PostgreSQL создаст свой
                    if 'id' in data:
                        del data['id']
                    
                    # Сохранение в PostgreSQL
                    client_id, saved_data = self.postgres_db.save_client_data_with_generated_id(data)
                    migrated_count += 1
                    
                    if migrated_count % 100 == 0:
                        print(f"Перенесено {migrated_count} записей...")
                
                except Exception as e:
                    print(f"Ошибка миграции записи {record.get('client_id', 'unknown')}: {e}")
                    error_count += 1
                    continue
            
            sqlite_conn.close()
            
            print(f"Миграция завершена:")
            print(f"Успешно перенесено: {migrated_count}")
            print(f"Ошибок: {error_count}")
            
            return migrated_count, error_count
            
        except Exception as e:
            print(f"Критическая ошибка миграции: {e}")
            raise e


def migrate_from_sqlite_to_postgresql(sqlite_path, postgres_params):
    """Функция для миграции данных из SQLite в PostgreSQL"""
    migration = MigrationManager(sqlite_path, postgres_params)
    return migration.migrate_from_sqlite()

def search_my_clients_by_fio_in_db(search_term, user_id, connection_params=None):
    """Поиск клиентов по ФИО только для конкретного пользователя с учетом ё/е"""
    try:
        db = DatabaseManager(connection_params)
        with db.get_connection() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                search_term = search_term.strip()
                results = []
                
                print(f"Поиск своих клиентов по ФИО: '{search_term}', user_id: {user_id}")
                
                # Функция для замены ё на е и наоборот
                def get_e_yo_variants(text):
                    variants = set()
                    variants.add(text)  # оригинал
                    
                    # Замена ё на е
                    if 'ё' in text.lower():
                        variants.add(text.replace('ё', 'е').replace('Ё', 'Е'))
                    
                    # Замена е на ё
                    if 'е' in text.lower():
                        variants.add(text.replace('е', 'ё').replace('Е', 'Ё'))
                    
                    return list(variants)
                
                # Генерируем варианты с учетом ё/е для поискового термина
                search_variants = get_e_yo_variants(search_term)
                print(f"Варианты поиска с ё/е: {search_variants}")
                
                # 1. Точное совпадение (с учетом ё/е)
                exact_patterns = set()
                for variant in search_variants:
                    exact_patterns.add(variant)
                    exact_patterns.add(variant.lower())
                    exact_patterns.add(variant.upper())
                    exact_patterns.add(variant.title())
                
                for pattern in exact_patterns:
                    query = '''
                    SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                           COALESCE(data_json, '{}') as data_json
                    FROM clients 
                    WHERE fio = %s AND agent_id = %s::text
                    ORDER BY id DESC
                    '''
                    
                    cursor.execute(query, (pattern, str(user_id)))
                    exact_results = cursor.fetchall()
                    if exact_results:
                        results.extend(exact_results)
                        print(f"Найдено точных совпадений для '{pattern}': {len(exact_results)}")
                
                # 2. Частичное совпадение (с учетом ё/е)
                if not results:
                    partial_patterns = set()
                    for variant in search_variants:
                        partial_patterns.add(f"%{variant}%")
                        partial_patterns.add(f"%{variant.lower()}%")
                        partial_patterns.add(f"%{variant.upper()}%")
                        partial_patterns.add(f"%{variant.title()}%")
                    
                    for pattern in partial_patterns:
                        query = '''
                        SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                               COALESCE(data_json, '{}') as data_json
                        FROM clients 
                        WHERE fio ILIKE %s AND agent_id = %s::text
                        ORDER BY id DESC
                        '''
                        
                        cursor.execute(query, (pattern, user_id))
                        partial_results = cursor.fetchall()

                        print(f"DEBUG SEARCH: pattern={pattern}, user_id={user_id}")
                        print(f"DEBUG SEARCH: found {len(partial_results)} results")
                        if partial_results:
                            results.extend(partial_results)
                            print(f"Найдено частичных совпадений для '{pattern}': {len(partial_results)}")
                
                # 3. Поиск по отдельным словам (с учетом ё/е)
                if not results:
                    search_words = search_term.split()
                    if len(search_words) >= 2:
                        first_word = search_words[0].strip()
                        second_word = search_words[1].strip()
                        
                        # Варианты с ё/е для каждого слова
                        first_word_variants = get_e_yo_variants(first_word)
                        second_word_variants = get_e_yo_variants(second_word)
                        
                        # Пробуем все комбинации
                        for first_variant in first_word_variants:
                            for second_variant in second_word_variants:
                                # Различные варианты регистра
                                for first_case in [first_variant, first_variant.lower(), first_variant.upper(), first_variant.title()]:
                                    for second_case in [second_variant, second_variant.lower(), second_variant.upper(), second_variant.title()]:
                                        query = '''
                                        SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                                               COALESCE(data_json, '{}') as data_json
                                        FROM clients 
                                        WHERE fio ILIKE %s AND fio ILIKE %s AND agent_id = %s::text
                                        ORDER BY id DESC
                                        '''
                                        
                                        cursor.execute(query, (f"%{first_case}%", f"%{second_case}%", user_id))
                                        word_results = cursor.fetchall()
                                        if word_results:
                                            results.extend(word_results)
                                            print(f"Найдено по словам '{first_case}' + '{second_case}': {len(word_results)}")
                                            break
                            
                            if results:
                                break
                
                # 4. Поиск только по первому слову (фамилии) с учетом ё/е
                if not results:
                    first_word = search_term.split()[0] if search_term.split() else search_term
                    first_word_variants = get_e_yo_variants(first_word)
                    
                    for variant in first_word_variants:
                        # Различные варианты регистра
                        for case_variant in [variant, variant.lower(), variant.upper(), variant.title()]:
                            query = '''
                            SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                                   COALESCE(data_json, '{}') as data_json
                            FROM clients 
                            WHERE fio ILIKE %s AND agent_id = %s::text
                            ORDER BY id DESC
                            '''
                            
                            cursor.execute(query, (f"%{case_variant}%", user_id))
                            surname_results = cursor.fetchall()
                            if surname_results:
                                results.extend(surname_results)
                                print(f"Найдено по фамилии '{case_variant}': {len(surname_results)}")
                                break
                
                # Удаляем дубликаты по client_id
                unique_results = []
                seen_client_ids = set()
                
                for result in results:
                    client_id = result['client_id']
                    if client_id not in seen_client_ids:
                        unique_results.append(dict(result))
                        seen_client_ids.add(client_id)
                
                print(f"Уникальных результатов поиска своих клиентов: {len(unique_results)}")
                
                return unique_results
    except Exception as e:
        print(f"Ошибка поиска своих клиентов по ФИО: {e}")
        return []

def search_city_clients_by_fio_in_db(search_term, admin_user_id, connection_params=None):
    """Поиск клиентов по ФИО в городе администратора"""
    try:
        db = DatabaseManager(connection_params)
        with db.get_connection() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
                search_term = search_term.strip()
                
                # Сначала получаем город администратора
                cursor.execute("SELECT city_admin FROM admins WHERE user_id = %s::text AND is_active = true", (admin_user_id,))
                admin_result = cursor.fetchone()
                
                if not admin_result or not admin_result['city_admin']:
                    print(f"Администратор {admin_user_id} не найден или не указан город")
                    return []
                
                admin_city = admin_result['city_admin']
                print(f"Поиск клиентов по ФИО: '{search_term}' в городе: '{admin_city}'")
                
                results = []
                
                # 1. Точное совпадение
                exact_patterns = [
                    search_term,
                    search_term.lower(),
                    search_term.upper(),
                    search_term.title()
                ]
                
                for pattern in exact_patterns:
                    query = '''
                    SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                           COALESCE(data_json, '{}') as data_json
                    FROM clients 
                    WHERE fio = %s AND city = %s
                    ORDER BY id DESC
                    '''
                    
                    cursor.execute(query, (pattern, admin_city))
                    exact_results = cursor.fetchall()
                    if exact_results:
                        results.extend(exact_results)
                        print(f"Найдено точных совпадений: {len(exact_results)}")
                        break
                
                # 2. Частичное совпадение
                if not results:
                    partial_patterns = [
                        f"%{search_term}%",
                        f"%{search_term.lower()}%", 
                        f"%{search_term.upper()}%",
                        f"%{search_term.title()}%"
                    ]
                    
                    for pattern in partial_patterns:
                        query = '''
                        SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                               COALESCE(data_json, '{}') as data_json
                        FROM clients 
                        WHERE fio ILIKE %s AND city = %s
                        ORDER BY id DESC
                        '''
                        
                        cursor.execute(query, (pattern, admin_city))
                        partial_results = cursor.fetchall()
                        if partial_results:
                            results.extend(partial_results)
                            print(f"Найдено частичных совпадений: {len(partial_results)}")
                            break
                
                # 3. Поиск по отдельным словам
                if not results:
                    search_words = search_term.split()
                    if len(search_words) >= 2:
                        first_word = search_words[0].strip()
                        second_word = search_words[1].strip()
                        
                        word_variants = []
                        for word in [first_word, second_word]:
                            word_variants.append([
                                word,
                                word.lower(),
                                word.upper(),
                                word.title()
                            ])
                        
                        for first_variants in word_variants[0]:
                            for second_variants in word_variants[1]:
                                query = '''
                                SELECT id, client_id, fio, number, car_number, date_dtp, created_at, 
                                       COALESCE(data_json, '{}') as data_json
                                FROM clients 
                                WHERE fio ILIKE %s AND fio ILIKE %s AND city = %s
                                ORDER BY id DESC
                                '''
                                
                                cursor.execute(query, (f"%{first_variants}%", f"%{second_variants}%", admin_city))
                                word_results = cursor.fetchall()
                                if word_results:
                                    results.extend(word_results)
                                    print(f"Найдено по словам '{first_variants}' + '{second_variants}': {len(word_results)}")
                                    break
                            
                            if results:
                                break
                
                # Удаляем дубликаты по client_id
                unique_results = []
                seen_client_ids = set()
                
                for result in results:
                    client_id = result['client_id']
                    if client_id not in seen_client_ids:
                        unique_results.append(dict(result))
                        seen_client_ids.add(client_id)
                
                print(f"Уникальных результатов поиска по городу '{admin_city}': {len(unique_results)}")
                
                return unique_results
    except Exception as e:
        print(f"Ошибка поиска клиентов по городу: {e}")
        return []
def get_agent_fio_by_id(agent_id):
    """Получить ФИО агента по его ID"""
    db = DatabaseManager()
    with db.get_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT fio FROM admins WHERE user_id = %s", (agent_id,))
            result = cursor.fetchone()

            return result[0] if result else "Неизвестный агент"



